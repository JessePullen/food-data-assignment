import openpyxl
import sqlite3
#Database connection
connection = sqlite3.connect("food.db")
cursor = connection.cursor()

#Open excel files and assign variable names to workbooks and sheets
wbV = openpyxl.load_workbook('violations.xlsx')
wbI = openpyxl.load_workbook('inspections.xlsx')
wsV = wbV['violations']
wsI = wbI['inspections']


#Create violations table
violationsTable = """
CREATE TABLE IF NOT EXISTS violations (
points INTEGER,
serial_number VARCHAR(10),
violation_code CHAR(4),
violation_description VARCHAR(100),
violation_status VARCHAR(20)
);"""

#Create inspections table
inspectionsTable = """
CREATE TABLE IF NOT EXISTS inspections (
activity_date DATE,
employee_id CHAR(9),
facility_address VARCHAR(30),
facility_city VARCHAR(30),
facility_id CHAR(9),
facility_name VARCHAR(30),
facility_state CHAR(2),
facility_zip CHAR(5),
grade CHAR(1),
owner_id CHAR(9),
owner_name VARCHAR(50),
pe_description VARCHAR(100),
program_element_pe CHAR(4),
program_name VARCHAR(30),
program_status VARCHAR(8),
record_id CHAR(9),
score VARCHAR(3),
serial_number CHAR(9) PRIMARY KEY,
service_code VARCHAR(3),
service_description VARCHAR(100)
);"""

#Execute create table commands
cursor.execute(violationsTable)
cursor.execute(inspectionsTable)

#Insert data commands for the two spreadsheets to the two tables
wsVCommand = """INSERT INTO violations (points, serial_number, violation_code, violation_description, violation_status)
VALUES (?, ?, ?, ?, ?) """
wsICommand = """INSERT INTO inspections (activity_date, employee_id, facility_address, facility_city, facility_id, facility_name, 
facility_state, facility_zip, grade, owner_id, owner_name, pe_description, program_element_pe, program_name, program_status,
record_id, score, serial_number, service_code, service_description)
VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) """

#Iterate through rows in the violations spreadsheet and assign values to columns
for row in wsV.iter_rows(min_row = 2):
    points                  = row[0]
    serial_number           = row[1]
    violation_code          = row[2]
    violation_description   = row[3]
    violation_status        = row[4]
    
    valuesV = (points.value, serial_number.value, violation_code.value, violation_description.value, violation_status.value)
    #execute SQL command
    cursor.execute(wsVCommand, valuesV)

#Iterale through rows in the violations spreadsheet and assign values to columns
for row in wsI.iter_rows(min_row = 2):
    activity_date           = row[0]
    employee_id             = row[1]
    facility_address        = row[2]
    facility_city           = row[3]
    facility_id             = row[4]
    facility_name           = row[5]
    facility_state          = row[6]
    facility_zip            = row[7]
    grade                   = row[8]
    owner_id                = row[9]
    owner_name              = row[10]
    pe_description          = row[11]
    program_element_pe      = row[12]
    program_name            = row[13]
    program_status          = row[14]
    record_id               = row[15]
    score                   = row[16]
    serial_number           = row[17]
    service_code            = row[18]
    service_description     = row[19]

    valuesI = (activity_date.value, employee_id.value, facility_address.value, facility_city.value, facility_id.value, facility_name.value, 
               facility_state.value, facility_zip.value, grade.value, owner_id.value, owner_name.value, pe_description.value, program_element_pe.value, 
               program_name.value, program_status.value, record_id.value, score.value, serial_number.value, service_code.value, service_description.value)
    #execute SQL command
    cursor.execute(wsICommand, valuesI)


print('Finished')    

#Commit and close database
connection.commit()
connection.close() 










